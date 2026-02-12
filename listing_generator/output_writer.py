"""
OUTPUT WRITER
=============
Writes the final listing data to Excel with required columns:
  date, la-cat, client_id, Country, rcm title(200), ai descr(300),
  rcm kf1-5(200 each), descrp(800-1500), search terms(200),
  main_image, lifestyle_image, why_choose_us_image

Images are embedded directly into the Excel cells.
"""

from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd


def _truncate(text: str, max_len: int) -> str:
    """Truncate text to max length, preserving whole words where possible."""
    text = str(text or "").strip()
    if len(text) <= max_len:
        return text
    cut = text[:max_len]
    last_space = cut.rfind(' ')
    if last_space > max_len * 0.7:
        return cut[:last_space].rstrip()
    return cut


def build_output_row(
    product: Dict[str, Any],
    optimized_title: str,
    ai_description: str,
    bullets: List[str],
    description: str,
    search_terms: str,
    image_paths: Optional[Dict[str, str]] = None,
) -> Dict[str, Any]:
    """
    Build a single output row conforming to the required schema.

    Columns: date, la-cat, client_id, Country, rcm title, ai descr,
             rcm kf1 - rcm kf5, descrp, search terms,
             main_image, lifestyle_image, why_choose_us_image
    """
    # Ensure bullets is always 5 items
    bp = list(bullets or [])
    while len(bp) < 5:
        bp.append("")

    img = image_paths or {}

    row: Dict[str, Any] = {
        "date": datetime.now().strftime("%Y-%m-%d"),
        "ASIN": product.get("asin", ""),
        "la-cat": product.get("la_cat", ""),
        "client_id": product.get("asin", ""),
        "Country": product.get("country", ""),
        "original_title": _truncate(product.get("title", ""), 300),
        "rcm title": _truncate(optimized_title, 200),
        "ai descr": _truncate(ai_description, 300),
        "rcm kf1": _truncate(bp[0], 200),
        "rcm kf2": _truncate(bp[1], 200),
        "rcm kf3": _truncate(bp[2], 200),
        "rcm kf4": _truncate(bp[3], 200),
        "rcm kf5": _truncate(bp[4], 200),
        "descrp": _truncate(description, 1500),
        "search terms": search_terms,
        "main_image": img.get("main_image", ""),
        "lifestyle_image_1": img.get("lifestyle_1", ""),
        "lifestyle_image_2": img.get("lifestyle_2", ""),
        "lifestyle_image_3": img.get("lifestyle_3", ""),
        "lifestyle_image_4": img.get("lifestyle_4", ""),
        "why_choose_us_image": img.get("why_choose_us", ""),
    }
    return row


def write_excel(
    rows: List[Dict[str, Any]],
    output_path: str,
) -> str:
    """
    Write all listing rows to an Excel file with images embedded in cells.
    Returns the path to the written file.
    """
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.utils import get_column_letter

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    # Column order (text columns first, then image columns)
    text_cols = [
        "date", "ASIN", "la-cat", "client_id", "Country",
        "original_title", "rcm title", "ai descr",
        "rcm kf1", "rcm kf2", "rcm kf3", "rcm kf4", "rcm kf5",
        "descrp", "search terms",
    ]
    image_cols = [
        "main_image",
        "lifestyle_image_1", "lifestyle_image_2", "lifestyle_image_3", "lifestyle_image_4",
        "why_choose_us_image"
    ]
    all_cols = text_cols + image_cols

    wb = Workbook()
    ws = wb.active
    ws.title = "Listings"

    # Write header
    for col_idx, col_name in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = cell.font.copy(bold=True)

    # Set image column widths (wider for images)
    for col_idx, col_name in enumerate(all_cols, 1):
        col_letter = get_column_letter(col_idx)
        if col_name in image_cols:
            ws.column_dimensions[col_letter].width = 22  # ~160px
        elif col_name in ("rcm title", "original_title", "descrp"):
            ws.column_dimensions[col_letter].width = 50
        elif col_name in ("ai descr", "search terms"):
            ws.column_dimensions[col_letter].width = 40
        elif col_name.startswith("rcm kf"):
            ws.column_dimensions[col_letter].width = 35
        else:
            ws.column_dimensions[col_letter].width = 15

    # Write rows with images
    IMAGE_HEIGHT = 120  # pixels
    IMAGE_WIDTH = 120   # pixels

    for row_idx, row_data in enumerate(rows, 2):
        has_image = False

        for col_idx, col_name in enumerate(all_cols, 1):
            value = row_data.get(col_name, "")

            if col_name in image_cols and value and os.path.isfile(str(value)):
                # Embed the image
                try:
                    img = XlImage(str(value))
                    img.width = IMAGE_WIDTH
                    img.height = IMAGE_HEIGHT
                    col_letter = get_column_letter(col_idx)
                    anchor = f"{col_letter}{row_idx}"
                    ws.add_image(img, anchor)
                    has_image = True
                except Exception as e:
                    # Fall back to file path text
                    ws.cell(row=row_idx, column=col_idx, value=str(value))
            else:
                ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")

        # Set row height if images are present
        if has_image:
            ws.row_dimensions[row_idx].height = 95  # ~120px

    wb.save(str(out))
    print(f"\n   📄 Output Excel: {out}")
    print(f"      Rows: {len(rows)}")
    print(f"      Images embedded: {sum(1 for r in rows if any(r.get(c) and os.path.isfile(str(r.get(c, ''))) for c in image_cols))}")
    return str(out)


def save_product_images(
    product: Dict[str, Any],
    image_results: Dict[str, bool],
    image_source_dir: str,
    output_base: str,
) -> str:
    """
    Organize generated images into product output folder.
    Returns path to the product image folder.
    """
    asin = product.get("asin", "PRODUCT")
    safe_name = "".join(c if c.isalnum() or c in ("-", "_") else "_" for c in asin)

    product_dir = Path(output_base) / "images" / safe_name
    product_dir.mkdir(parents=True, exist_ok=True)

    # Copy/move generated images if they exist in the source
    source = Path(image_source_dir) if image_source_dir else None
    if source and source.exists():
        files_to_copy = [
            "main_product.png",
            "lifestyle_1.png",
            "lifestyle_2.png",
            "lifestyle_3.png",
            "lifestyle_4.png",
            "why_choose_us.png"
        ]
        for img_name in files_to_copy:
            src_file = source / img_name
            if src_file.exists():
                dst_file = product_dir / img_name
                if not dst_file.exists():
                    import shutil
                    shutil.copy2(str(src_file), str(dst_file))

    return str(product_dir)


def write_analysis_json(
    product: Dict[str, Any],
    image_analysis: Dict[str, Any],
    optimized_title: str,
    output_dir: str,
) -> str:
    """Save debug/analysis JSON for a product."""
    import json

    asin = product.get("asin", "PRODUCT")
    safe_name = "".join(c if c.isalnum() or c in ("-", "_") else "_" for c in asin)

    analysis_dir = Path(output_dir) / "analysis"
    analysis_dir.mkdir(parents=True, exist_ok=True)

    data = {
        "asin": asin,
        "original_title": product.get("title", ""),
        "optimized_title": optimized_title,
        "country": product.get("country", ""),
        "la_cat": product.get("la_cat", ""),
        "image_analysis": {
            k: v for k, v in (image_analysis or {}).items()
            if k not in ("local_image_paths",)  # skip large fields
        },
        "timestamp": datetime.now().isoformat(),
    }

    path = analysis_dir / f"{safe_name}_analysis.json"
    with open(path, "w") as f:
        json.dump(data, f, indent=2, default=str)

    return str(path)


def load_existing_excel(output_dir: str) -> List[Dict[str, Any]]:
    """Load existing listing_output.xlsx rows so they can be preserved on resume.

    Reads text columns via pandas AND reconstructs image file paths from
    the images/ folder so they get re-embedded when the Excel is rewritten.
    """
    excel_path = os.path.join(output_dir, "listing_output.xlsx")
    if not os.path.isfile(excel_path):
        return []

    try:
        df = pd.read_excel(excel_path, sheet_name="Listings")
    except Exception as e:
        print(f"   ⚠️ Could not read existing Excel: {e}")
        return []

    text_cols = [
        "date", "ASIN", "la-cat", "client_id", "Country",
        "original_title", "rcm title", "ai descr",
        "rcm kf1", "rcm kf2", "rcm kf3", "rcm kf4", "rcm kf5",
        "descrp", "search terms",
    ]

    # Image file names mapping
    image_file_map = {
        "main_image": "main_product.png",
        "lifestyle_image_1": "lifestyle_1.png",
        "lifestyle_image_2": "lifestyle_2.png",
        "lifestyle_image_3": "lifestyle_3.png",
        "lifestyle_image_4": "lifestyle_4.png",
        "why_choose_us_image": "why_choose_us.png",
    }

    # Only read columns that actually exist in the file
    existing_cols = set(df.columns)

    rows: List[Dict[str, Any]] = []
    for row_idx, df_row in df.iterrows():
        row: Dict[str, Any] = {}
        for col in text_cols:
            if col in existing_cols:
                val = df_row.get(col, "")
                row[col] = str(val) if pd.notna(val) else ""
            else:
                row[col] = ""

        # If old Excel has no ASIN column, copy from client_id
        if not row.get("ASIN") and row.get("client_id"):
            row["ASIN"] = row["client_id"]

        # Reconstruct image paths from images/product_{row_idx}/ folder
        img_dir = os.path.join(output_dir, "images", f"product_{row_idx}")
        for col_name, file_name in image_file_map.items():
            fpath = os.path.join(img_dir, file_name)
            row[col_name] = fpath if os.path.isfile(fpath) else ""

        rows.append(row)

    img_count = sum(1 for r in rows if any(r.get(c) for c in image_file_map))
    print(f"   📂 Read {len(rows)} existing rows from {excel_path} ({img_count} with images)")
    return rows
